#!/usr/bin/env python3
"""
Excel export functionality for JIRA metrics reports.
Creates Excel files with multiple sheets optimized for pivot tables and charts.
"""

import pandas as pd
from datetime import datetime
from typing import Dict, List
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Export JIRA metrics reports to Excel format."""

    def __init__(self, report: Dict):
        """Initialize with a report dictionary."""
        self.report = report

    def export_to_excel(self, filename: str) -> str:
        """
        Export report to Excel file with multiple sheets.

        Args:
            filename: Output Excel filename

        Returns:
            Path to the created Excel file
        """
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Sheet 1: Weekly Summary (for time series charts)
            self._export_weekly_summary(writer)

            # Sheet 2: Overall Metrics Summary
            self._export_overall_summary(writer)

            # Sheet 3: Team Comparison
            if self.report.get("by_team"):
                self._export_team_comparison(writer)

            # Sheet 4: Individual Performance
            self._export_individual_performance(writer)

            # Sheet 5: Raw Issue Data (for pivot tables)
            self._export_raw_issues(writer)

        # Apply formatting
        self._apply_formatting(filename)

        return filename

    def _export_weekly_summary(self, writer):
        """Export weekly summary data for time series analysis."""
        weekly_data = self.report["overall"]["weekly_summary"]["weeks"]

        rows = []
        for week, data in sorted(weekly_data.items()):
            row = {
                "Week": week,
                "Issues Completed": data["issue_count"],
                "Cycle Time Mean (days)": data["cycle_time"]["mean_days"],
                "Cycle Time P50 (days)": data["cycle_time"]["median_days"],
                "Cycle Time P85 (days)": data["cycle_time"]["p85_days"],
                "Lead Time Mean (days)": data["lead_time"]["mean_days"],
                "Lead Time P50 (days)": data["lead_time"]["median_days"],
                "Lead Time P85 (days)": data["lead_time"]["p85_days"],
            }
            rows.append(row)

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Weekly Summary", index=False)

    def _export_overall_summary(self, writer):
        """Export overall summary metrics."""
        overall_cycle = self.report["overall"]["cycle_time"]
        overall_lead = self.report["overall"]["lead_time"]

        data = {
            "Metric": [
                "Total Issues",
                "Cycle Time - Mean (days)",
                "Cycle Time - P50 (days)",
                "Cycle Time - P85 (days)",
                "Cycle Time - P95 (days)",
                "Lead Time - Mean (days)",
                "Lead Time - P50 (days)",
                "Lead Time - P85 (days)",
                "Lead Time - P95 (days)",
            ],
            "Value": [
                overall_cycle["sample_size"],
                overall_cycle["mean_cycle_time_days"],
                overall_cycle["median_cycle_time_days"],
                overall_cycle["p85_cycle_time_days"],
                overall_cycle["p95_cycle_time_days"],
                overall_lead["mean_lead_time_days"],
                overall_lead["median_lead_time_days"],
                overall_lead["p85_lead_time_days"],
                overall_lead["p95_lead_time_days"],
            ]
        }

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="Overall Summary", index=False)

    def _export_team_comparison(self, writer):
        """Export team comparison data for charts."""
        rows = []

        for team, team_data in self.report["by_team"].items():
            team_cycle = team_data["cycle_time"]
            team_lead = team_data["lead_time"]

            row = {
                "Team": team,
                "Issues Completed": team_cycle["sample_size"],
                "Cycle Time Mean (days)": team_cycle["mean_cycle_time_days"],
                "Cycle Time P50 (days)": team_cycle["median_cycle_time_days"],
                "Cycle Time P85 (days)": team_cycle["p85_cycle_time_days"],
                "Lead Time Mean (days)": team_lead["mean_lead_time_days"],
                "Lead Time P50 (days)": team_lead["median_lead_time_days"],
                "Lead Time P85 (days)": team_lead["p85_lead_time_days"],
            }
            rows.append(row)

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Team Comparison", index=False)

    def _export_individual_performance(self, writer):
        """Export individual team member performance."""
        rows = []

        # If we have team breakdown
        if self.report.get("by_team"):
            for team, team_data in self.report["by_team"].items():
                for member, member_data in team_data.get("members", {}).items():
                    member_cycle = member_data["cycle_time"]
                    member_lead = member_data["lead_time"]

                    row = {
                        "Team": team,
                        "Team Member": member,
                        "Issues Completed": member_cycle["sample_size"],
                        "Cycle Time Mean (days)": member_cycle["mean_cycle_time_days"],
                        "Cycle Time P50 (days)": member_cycle["median_cycle_time_days"],
                        "Cycle Time P85 (days)": member_cycle["p85_cycle_time_days"],
                        "Lead Time Mean (days)": member_lead["mean_lead_time_days"],
                        "Lead Time P50 (days)": member_lead["median_lead_time_days"],
                        "Lead Time P85 (days)": member_lead["p85_lead_time_days"],
                    }
                    rows.append(row)
        else:
            # Individual contributors without team breakdown
            for member, member_data in self.report.get("by_team_member", {}).items():
                member_cycle = member_data["cycle_time"]
                member_lead = member_data["lead_time"]

                row = {
                    "Team": "All",
                    "Team Member": member,
                    "Issues Completed": member_cycle["sample_size"],
                    "Cycle Time Mean (days)": member_cycle["mean_cycle_time_days"],
                    "Cycle Time P50 (days)": member_cycle["median_cycle_time_days"],
                    "Cycle Time P85 (days)": member_cycle["p85_cycle_time_days"],
                    "Lead Time Mean (days)": member_lead["mean_lead_time_days"],
                    "Lead Time P50 (days)": member_lead["median_lead_time_days"],
                    "Lead Time P85 (days)": member_lead["p85_lead_time_days"],
                }
                rows.append(row)

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Individual Performance", index=False)

    def _export_raw_issues(self, writer):
        """Export raw issue data for pivot tables."""
        rows = []

        # Collect all issues from overall metrics
        overall_cycle = self.report["overall"]["cycle_time"]
        overall_lead = self.report["overall"]["lead_time"]

        # Create a map of lead time by issue key
        lead_time_map = {}
        for issue in overall_lead.get("issues", []):
            lead_time_map[issue["key"]] = {
                "lead_time_days": issue["lead_time_days"],
                "lead_time_hours": issue["lead_time_hours"]
            }

        for issue in overall_cycle.get("issues", []):
            # Extract project from issue key (e.g., "POD1-1068" -> "POD1")
            issue_key = issue["key"]
            project = issue_key.split("-")[0] if "-" in issue_key else "Unknown"

            # Get lead time for this issue
            lead_time_data = lead_time_map.get(issue_key, {})

            row = {
                "Issue Key": issue_key,
                "Project": project,
                "Summary": issue["summary"],
                "Assignee": issue["assignee"],
                "Resolved Date": issue["resolved_date"],
                "Cycle Time (days)": issue["cycle_time_days"],
                "Cycle Time (hours)": issue["cycle_time_hours"],
                "Lead Time (days)": lead_time_data.get("lead_time_days", issue["cycle_time_days"]),
                "Lead Time (hours)": lead_time_data.get("lead_time_hours", issue["cycle_time_hours"]),
            }
            rows.append(row)

        df = pd.DataFrame(rows)

        # Add week column for grouping
        if not df.empty:
            df["Resolved Date"] = pd.to_datetime(df["Resolved Date"])
            df["Year-Week"] = df["Resolved Date"].dt.strftime("%Y-W%V")
            df["Month"] = df["Resolved Date"].dt.strftime("%Y-%m")

            # Sort by resolved date
            df = df.sort_values("Resolved Date")

        df.to_excel(writer, sheet_name="Raw Issue Data", index=False)

    def _apply_formatting(self, filename: str):
        """Apply formatting to the Excel file."""
        wb = load_workbook(filename)

        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze top row
            ws.freeze_panes = "A2"

        wb.save(filename)


def export_report_to_excel(report: Dict, filename: str = None) -> str:
    """
    Convenience function to export a report to Excel.

    Args:
        report: Report dictionary from generate_team_performance_report()
        filename: Optional output filename (auto-generated if not provided)

    Returns:
        Path to created Excel file
    """
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"jira_metrics_{timestamp}.xlsx"

    exporter = ExcelExporter(report)
    return exporter.export_to_excel(filename)
