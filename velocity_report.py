#!/usr/bin/env python3
"""
Velocity & Ticket Count Report

Tracks developer velocity (story points/week) and ticket throughput over
a configurable period, broken down by developer with weekly granularity.

Usage:
    python velocity_report.py --jira-url https://yourco.atlassian.net \
        --email you@example.com --api-token TOKEN \
        --projects IA,DATA,SAOP,SAOP2 --start-date 2025-09-05 --end-date 2026-03-05
"""

import os
import csv
import json
import argparse
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Dict, List, Tuple, Optional
from dora_metrics import JiraDORAMetrics

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Common story point field names across Jira instances
STORY_POINT_FIELDS = [
    "customfield_10036",  # NINJIO's field
    "customfield_10016",
    "customfield_10004",
    "customfield_10002",
    "Story Points",
    "storyPoints",
]


class VelocityReport:
    """Generate velocity and ticket count reports per developer."""

    def __init__(self, jira_url: str, email: str, api_token: str):
        self.calculator = JiraDORAMetrics(jira_url, email, api_token)

    def fetch_completed_issues(self, projects: List[str], start_date: str, end_date: str) -> List[Dict]:
        """Fetch completed issues in the date range. No changelog needed."""
        jql = (
            f"project in ({','.join(projects)}) "
            f"AND resolutiondate >= '{start_date}' AND resolutiondate <= '{end_date}' "
            f"AND status in (Done, Resolved, Closed, Released, Deployed) "
            f"AND issuetype in (Story, Bug, Task)"
        )
        fields = ["summary", "assignee", "resolutiondate", "issuetype"] + STORY_POINT_FIELDS
        print(f"Fetching completed issues from {start_date} to {end_date}...")
        issues = self.calculator.search_issues(jql, fields=fields)
        print(f"  Found {len(issues)} issues.")
        return issues

    def _extract_story_points(self, issue_fields: Dict) -> Tuple[float, bool]:
        """
        Extract story points from issue fields.

        Returns:
            (points_value, has_story_points) - value is 0.0 if no SP found.
        """
        for field in STORY_POINT_FIELDS:
            value = issue_fields.get(field)
            if value is not None:
                try:
                    return (float(value), True)
                except (ValueError, TypeError):
                    continue
        return (0.0, False)

    def _get_week_start(self, date_str: str) -> str:
        """Return the Monday (ISO week start) for a given date string."""
        dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
        monday = dt - timedelta(days=dt.weekday())
        return monday.strftime("%Y-%m-%d")

    def _get_developer_name(self, assignee: Optional[Dict]) -> str:
        if assignee and assignee.get("displayName"):
            return assignee["displayName"]
        return "Unassigned"

    def group_by_developer_and_week(
        self, issues: List[Dict], start_date: str
    ) -> Dict[str, Dict[str, Dict]]:
        """
        Group issues into nested dict:
          developer -> week_start -> {ticket_count, story_points, tickets_with_sp, tickets_without_sp}
        """
        grouped: Dict[str, Dict[str, Dict]] = defaultdict(lambda: defaultdict(
            lambda: {"ticket_count": 0, "story_points": 0.0, "tickets_with_sp": 0, "tickets_without_sp": 0}
        ))

        for issue in issues:
            fields = issue.get("fields", {})
            developer = self._get_developer_name(fields.get("assignee"))
            resolution_date = fields.get("resolutiondate", "")
            if not resolution_date:
                continue

            week_start = self._get_week_start(resolution_date)
            points, has_sp = self._extract_story_points(fields)

            entry = grouped[developer][week_start]
            entry["ticket_count"] += 1
            if has_sp:
                entry["story_points"] += points
                entry["tickets_with_sp"] += 1
            else:
                entry["tickets_without_sp"] += 1

        return grouped

    def _generate_all_weeks(self, start_date: str, end_date: str) -> List[str]:
        """Generate all Monday-start week dates covering the period."""
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        # Align to Monday
        first_monday = start_dt - timedelta(days=start_dt.weekday())
        weeks = []
        current = first_monday
        while current <= end_dt:
            weeks.append(current.strftime("%Y-%m-%d"))
            current += timedelta(days=7)
        return weeks

    def calculate_developer_summaries(
        self, grouped_data: Dict[str, Dict[str, Dict]], all_weeks: List[str]
    ) -> List[Dict]:
        """Compute totals and averages per developer."""
        summaries = []
        for developer, week_data in sorted(grouped_data.items()):
            total_tickets = 0
            total_sp = 0.0
            total_with_sp = 0
            total_without_sp = 0
            active_weeks = 0

            for week in all_weeks:
                entry = week_data.get(week)
                if entry and entry["ticket_count"] > 0:
                    active_weeks += 1
                    total_tickets += entry["ticket_count"]
                    total_sp += entry["story_points"]
                    total_with_sp += entry["tickets_with_sp"]
                    total_without_sp += entry["tickets_without_sp"]

            weeks_for_avg = active_weeks if active_weeks > 0 else 1
            summaries.append({
                "developer": developer,
                "total_tickets": total_tickets,
                "total_story_points": total_sp,
                "tickets_with_sp": total_with_sp,
                "tickets_without_sp": total_without_sp,
                "avg_tickets_per_week": round(total_tickets / weeks_for_avg, 1),
                "avg_sp_per_week": round(total_sp / weeks_for_avg, 1),
                "weeks_active": active_weeks,
            })

        # Sort by total SP descending
        summaries.sort(key=lambda s: s["total_story_points"], reverse=True)
        return summaries

    def generate_velocity_report(
        self, projects: List[str], start_date: str, end_date: str
    ) -> Dict:
        """Main entry point: fetch, group, summarize."""
        issues = self.fetch_completed_issues(projects, start_date, end_date)
        all_weeks = self._generate_all_weeks(start_date, end_date)
        grouped = self.group_by_developer_and_week(issues, start_date)
        summaries = self.calculate_developer_summaries(grouped, all_weeks)

        return {
            "period": {"start_date": start_date, "end_date": end_date},
            "projects": projects,
            "total_weeks": len(all_weeks),
            "total_issues": len(issues),
            "developer_summaries": summaries,
            "weekly_breakdown": {
                dev: [
                    {
                        "week_start": w,
                        "week_end": (datetime.strptime(w, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d"),
                        **grouped[dev].get(w, {"ticket_count": 0, "story_points": 0.0, "tickets_with_sp": 0, "tickets_without_sp": 0}),
                    }
                    for w in all_weeks
                ]
                for dev in grouped
            },
        }

    def print_report(self, report: Dict):
        """Print formatted console output."""
        period = report["period"]
        projects = ", ".join(report["projects"])
        print()
        print("=" * 90)
        print("VELOCITY & TICKET COUNT REPORT")
        print(f"Period: {period['start_date']} to {period['end_date']} | Projects: {projects}")
        print(f"Total issues: {report['total_issues']} | Weeks in range: {report['total_weeks']}")
        print("=" * 90)

        # Developer summary table
        print()
        print("DEVELOPER SUMMARY")
        header = (
            f"{'Developer':<30} {'Total Tix':>10} {'Total SP':>10} "
            f"{'Avg Tix/Wk':>11} {'Avg SP/Wk':>10} {'Weeks Active':>13}"
        )
        print(header)
        print("-" * 90)

        for s in report["developer_summaries"]:
            print(
                f"{s['developer']:<30} {s['total_tickets']:>10} "
                f"{s['total_story_points']:>10.1f} {s['avg_tickets_per_week']:>11.1f} "
                f"{s['avg_sp_per_week']:>10.1f} {s['weeks_active']:>13}"
            )

        # Weekly breakdown per developer
        for s in report["developer_summaries"]:
            dev = s["developer"]
            weeks = report["weekly_breakdown"].get(dev, [])
            active = [w for w in weeks if w["ticket_count"] > 0]
            if not active:
                continue
            print()
            print(f"WEEKLY BREAKDOWN: {dev}")
            print(f"{'Week':<32} {'Tickets':>8} {'Story Points':>13} {'Tix w/SP':>9} {'Tix w/o SP':>11}")
            print("-" * 78)
            for w in active:
                label = f"{w['week_start']} to {w['week_end']}"
                print(
                    f"{label:<32} {w['ticket_count']:>8} "
                    f"{w['story_points']:>13.1f} {w['tickets_with_sp']:>9} "
                    f"{w['tickets_without_sp']:>11}"
                )

        print()
        print("=" * 90)

    def export_json(self, report: Dict, path: str):
        """Export report to JSON file."""
        with open(path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, default=str)
        print(f"JSON report saved to: {path}")

    def export_csv(self, report: Dict, path: str):
        """Export developer summary + weekly breakdown to CSV."""
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Developer summary section
            writer.writerow(["DEVELOPER SUMMARY"])
            writer.writerow([
                "Developer", "Total Tickets", "Total Story Points",
                "Avg Tickets/Week", "Avg SP/Week", "Weeks Active",
                "Tickets w/ SP", "Tickets w/o SP",
            ])
            for s in report["developer_summaries"]:
                writer.writerow([
                    s["developer"], s["total_tickets"], s["total_story_points"],
                    s["avg_tickets_per_week"], s["avg_sp_per_week"], s["weeks_active"],
                    s["tickets_with_sp"], s["tickets_without_sp"],
                ])

            writer.writerow([])

            # Weekly breakdown section
            writer.writerow(["WEEKLY BREAKDOWN"])
            writer.writerow([
                "Developer", "Week Start", "Week End",
                "Tickets", "Story Points", "Tix w/ SP", "Tix w/o SP",
            ])
            for dev, weeks in report["weekly_breakdown"].items():
                for w in weeks:
                    if w["ticket_count"] > 0:
                        writer.writerow([
                            dev, w["week_start"], w["week_end"],
                            w["ticket_count"], w["story_points"],
                            w["tickets_with_sp"], w["tickets_without_sp"],
                        ])

        print(f"CSV report saved to: {path}")

    def export_excel(self, report: Dict, path: str):
        """Export report to Excel with formatted tables and charts."""
        if not HAS_OPENPYXL:
            print("openpyxl not installed. Run: pip install openpyxl")
            return

        wb = Workbook()
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        thin_border = Border(
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # ── Sheet 1: Developer Summary ──────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Developer Summary"

        # Title row
        period = report["period"]
        projects = ", ".join(report["projects"])
        ws_summary.append([f"Velocity Report  |  {period['start_date']} to {period['end_date']}  |  {projects}"])
        ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws_summary["A1"].font = Font(bold=True, size=14, color="2F5496")
        ws_summary.append([])

        # Summary table headers
        summary_headers = [
            "Developer", "Total Tickets", "Total SP",
            "Avg Tix/Wk", "Avg SP/Wk", "Weeks Active",
            "Tix w/ SP", "Tix w/o SP",
        ]
        ws_summary.append(summary_headers)
        for col_idx, _ in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        summaries = report["developer_summaries"]
        for i, s in enumerate(summaries):
            row = [
                s["developer"], s["total_tickets"], s["total_story_points"],
                s["avg_tickets_per_week"], s["avg_sp_per_week"], s["weeks_active"],
                s["tickets_with_sp"], s["tickets_without_sp"],
            ]
            ws_summary.append(row)
            for col_idx in range(1, len(row) + 1):
                cell = ws_summary.cell(row=4 + i, column=col_idx)
                cell.border = thin_border
                if col_idx > 1:
                    cell.alignment = Alignment(horizontal="center")
                if i % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Column widths
        ws_summary.column_dimensions["A"].width = 32
        for col in "BCDEFGH":
            ws_summary.column_dimensions[col].width = 14

        # ── Chart 1: Total Tickets vs Total SP (bar comparison) ─────
        num_devs = len(summaries)
        data_start_row = 3  # header row
        data_end_row = 3 + num_devs

        chart_bar = BarChart()
        chart_bar.type = "col"
        chart_bar.title = "Total Tickets vs Story Points by Developer"
        chart_bar.y_axis.title = "Count"
        chart_bar.x_axis.title = "Developer"
        chart_bar.style = 10
        chart_bar.width = 28
        chart_bar.height = 14

        cats = Reference(ws_summary, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
        tickets_data = Reference(ws_summary, min_col=2, min_row=data_start_row, max_row=data_end_row)
        sp_data = Reference(ws_summary, min_col=3, min_row=data_start_row, max_row=data_end_row)
        chart_bar.add_data(tickets_data, titles_from_data=True)
        chart_bar.add_data(sp_data, titles_from_data=True)
        chart_bar.set_categories(cats)
        chart_bar.series[0].graphicalProperties.solidFill = "4472C4"
        chart_bar.series[1].graphicalProperties.solidFill = "ED7D31"

        ws_summary.add_chart(chart_bar, f"A{data_end_row + 3}")

        # ── Chart 2: Avg SP/Week comparison ─────────────────────────
        chart_avg = BarChart()
        chart_avg.type = "col"
        chart_avg.title = "Average Story Points per Week by Developer"
        chart_avg.y_axis.title = "Avg SP/Week"
        chart_avg.style = 10
        chart_avg.width = 28
        chart_avg.height = 14

        avg_sp_data = Reference(ws_summary, min_col=5, min_row=data_start_row, max_row=data_end_row)
        chart_avg.add_data(avg_sp_data, titles_from_data=True)
        chart_avg.set_categories(cats)
        chart_avg.series[0].graphicalProperties.solidFill = "70AD47"

        ws_summary.add_chart(chart_avg, f"A{data_end_row + 20}")

        # ── Sheet 2: Weekly Team Totals + trend chart ───────────────
        ws_weekly = wb.create_sheet("Weekly Totals")
        all_weeks = sorted({
            w["week_start"]
            for weeks in report["weekly_breakdown"].values()
            for w in weeks
            if w["ticket_count"] > 0
        })

        # Aggregate team totals per week
        week_totals = {}
        for week_key in all_weeks:
            week_totals[week_key] = {"tickets": 0, "story_points": 0.0, "active_devs": 0}
        for dev, weeks in report["weekly_breakdown"].items():
            for w in weeks:
                if w["week_start"] in week_totals and w["ticket_count"] > 0:
                    week_totals[w["week_start"]]["tickets"] += w["ticket_count"]
                    week_totals[w["week_start"]]["story_points"] += w["story_points"]
                    week_totals[w["week_start"]]["active_devs"] += 1

        ws_weekly.append(["Week", "Total Tickets", "Total Story Points", "Active Developers", "SP / Developer"])
        for col_idx in range(1, 6):
            cell = ws_weekly.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, wk in enumerate(all_weeks):
            t = week_totals[wk]
            sp_per_dev = round(t["story_points"] / t["active_devs"], 1) if t["active_devs"] > 0 else 0.0
            ws_weekly.append([wk, t["tickets"], t["story_points"], t["active_devs"], sp_per_dev])

        ws_weekly.column_dimensions["A"].width = 14
        ws_weekly.column_dimensions["B"].width = 16
        ws_weekly.column_dimensions["C"].width = 20
        ws_weekly.column_dimensions["D"].width = 18
        ws_weekly.column_dimensions["E"].width = 16

        num_weeks = len(all_weeks)
        chart_trend = LineChart()
        chart_trend.title = "Team Velocity Trend (Weekly)"
        chart_trend.y_axis.title = "Count"
        chart_trend.x_axis.title = "Week"
        chart_trend.style = 10
        chart_trend.width = 30
        chart_trend.height = 14

        cats_wk = Reference(ws_weekly, min_col=1, min_row=2, max_row=1 + num_weeks)
        tix_ref = Reference(ws_weekly, min_col=2, min_row=1, max_row=1 + num_weeks)
        sp_ref = Reference(ws_weekly, min_col=3, min_row=1, max_row=1 + num_weeks)
        devs_ref = Reference(ws_weekly, min_col=4, min_row=1, max_row=1 + num_weeks)
        chart_trend.add_data(tix_ref, titles_from_data=True)
        chart_trend.add_data(sp_ref, titles_from_data=True)
        chart_trend.add_data(devs_ref, titles_from_data=True)
        chart_trend.set_categories(cats_wk)
        chart_trend.series[0].graphicalProperties.line.solidFill = "4472C4"
        chart_trend.series[1].graphicalProperties.line.solidFill = "ED7D31"
        # Active Developers on secondary axis (different scale)
        chart_trend.series[2].graphicalProperties.line.solidFill = "70AD47"
        chart_trend.series[2].graphicalProperties.line.dashStyle = "dash"
        chart_trend.series[2].graphicalProperties.line.width = 25000
        from openpyxl.chart.axis import NumericAxis
        chart_trend.y_axis2 = NumericAxis(axId=200, crosses="max")
        chart_trend.y_axis2.title = "Active Developers"

        ws_weekly.add_chart(chart_trend, f"A{num_weeks + 4}")

        # ── Chart: SP per Developer efficiency trend ──────────────
        chart_eff = LineChart()
        chart_eff.title = "Efficiency: Story Points per Active Developer (Weekly)"
        chart_eff.y_axis.title = "SP / Developer"
        chart_eff.x_axis.title = "Week"
        chart_eff.style = 10
        chart_eff.width = 30
        chart_eff.height = 14

        sp_per_dev_ref = Reference(ws_weekly, min_col=5, min_row=1, max_row=1 + num_weeks)
        chart_eff.add_data(sp_per_dev_ref, titles_from_data=True)
        chart_eff.set_categories(cats_wk)
        chart_eff.series[0].graphicalProperties.line.solidFill = "7030A0"

        ws_weekly.add_chart(chart_eff, f"A{num_weeks + 21}")

        # ── Sheet 3: Per-Developer Weekly Breakdown ─────────────────
        ws_detail = wb.create_sheet("Developer Weekly Detail")
        detail_headers = [
            "Developer", "Week Start", "Week End",
            "Tickets", "Story Points", "Tix w/ SP", "Tix w/o SP",
        ]
        ws_detail.append(detail_headers)
        for col_idx in range(1, len(detail_headers) + 1):
            cell = ws_detail.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row_num = 2
        for s in summaries:
            dev = s["developer"]
            weeks = report["weekly_breakdown"].get(dev, [])
            for w in weeks:
                if w["ticket_count"] > 0:
                    ws_detail.append([
                        dev, w["week_start"], w["week_end"],
                        w["ticket_count"], w["story_points"],
                        w["tickets_with_sp"], w["tickets_without_sp"],
                    ])
                    row_num += 1

        ws_detail.column_dimensions["A"].width = 32
        for col in "BCDEFG":
            ws_detail.column_dimensions[col].width = 15

        # ── Sheet 4: Top Developer Trends (individual line charts) ──
        # Show weekly SP trend for top 5 developers by total SP
        top_devs = summaries[:5]
        ws_top = wb.create_sheet("Top Developer Trends")

        # Build a table: Week | Dev1 SP | Dev2 SP | ...
        top_dev_names = [s["developer"] for s in top_devs]
        ws_top.append(["Week"] + top_dev_names)
        for col_idx in range(1, len(top_dev_names) + 2):
            cell = ws_top.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for wk in all_weeks:
            row_data = [wk]
            for dev_name in top_dev_names:
                dev_weeks = report["weekly_breakdown"].get(dev_name, [])
                sp = 0.0
                for w in dev_weeks:
                    if w["week_start"] == wk:
                        sp = w["story_points"]
                        break
                row_data.append(sp)
            ws_top.append(row_data)

        ws_top.column_dimensions["A"].width = 14
        for col_idx in range(2, len(top_dev_names) + 2):
            ws_top.column_dimensions[get_column_letter(col_idx)].width = 20

        chart_devs = LineChart()
        chart_devs.title = "Top 5 Developers - Weekly Story Points"
        chart_devs.y_axis.title = "Story Points"
        chart_devs.x_axis.title = "Week"
        chart_devs.style = 10
        chart_devs.width = 30
        chart_devs.height = 16

        cats_top = Reference(ws_top, min_col=1, min_row=2, max_row=1 + num_weeks)
        for col_idx in range(2, len(top_dev_names) + 2):
            data_ref = Reference(ws_top, min_col=col_idx, min_row=1, max_row=1 + num_weeks)
            chart_devs.add_data(data_ref, titles_from_data=True)
        chart_devs.set_categories(cats_top)

        colors = ["4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5"]
        for i, series in enumerate(chart_devs.series):
            series.graphicalProperties.line.solidFill = colors[i % len(colors)]

        ws_top.add_chart(chart_devs, f"A{num_weeks + 4}")

        wb.save(path)
        print(f"Excel report saved to: {path}")


def main():
    parser = argparse.ArgumentParser(description="Generate velocity & ticket count report")
    parser.add_argument("--jira-url", required=True,
                        help="JIRA instance URL (e.g., https://yourcompany.atlassian.net)")
    parser.add_argument("--email", required=True, help="Atlassian account email")
    parser.add_argument("--api-token", required=True, help="Atlassian API token")
    parser.add_argument("--projects", default="IA,DATA,SAOP,SAOP2",
                        help="Comma-separated project keys (default: IA,DATA,SAOP,SAOP2)")
    parser.add_argument("--start-date",
                        default=(datetime.now() - timedelta(weeks=26)).strftime("%Y-%m-%d"),
                        help="Start date YYYY-MM-DD (default: 26 weeks ago)")
    parser.add_argument("--end-date",
                        default=datetime.now().strftime("%Y-%m-%d"),
                        help="End date YYYY-MM-DD (default: today)")
    parser.add_argument("--output", help="Output file path (without extension)")
    parser.add_argument("--format", choices=["json", "csv", "xlsx", "all"], default="all",
                        help="Export format (default: all)")
    parser.add_argument("--from-json",
                        help="Load report data from a previous JSON export instead of querying Jira")

    args = parser.parse_args()
    projects = [p.strip() for p in args.projects.split(",")]

    if args.from_json:
        with open(args.from_json, "r", encoding="utf-8") as f:
            report = json.load(f)
        print(f"Loaded report from {args.from_json}")
        report_gen = VelocityReport.__new__(VelocityReport)
    else:
        report_gen = VelocityReport(args.jira_url, args.email, args.api_token)
        report = report_gen.generate_velocity_report(projects, args.start_date, args.end_date)

    report_gen.print_report(report)

    if args.output:
        base = args.output
        fmt = args.format
        if fmt in ("json", "all"):
            report_gen.export_json(report, f"{base}.json")
        if fmt in ("csv", "all"):
            report_gen.export_csv(report, f"{base}.csv")
        if fmt in ("xlsx", "all"):
            report_gen.export_excel(report, f"{base}.xlsx")


if __name__ == "__main__":
    main()
